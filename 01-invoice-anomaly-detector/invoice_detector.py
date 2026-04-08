"""

  Invoice Anomaly Detector
  Consulting Portfolio Project | IT-Controlling / Finance
  """


import os
import random
import warnings
from datetime import datetime, timedelta

import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from sklearn.ensemble import IsolationForest
from sklearn.preprocessing import LabelEncoder

warnings.filterwarnings("ignore")
random.seed(42)
np.random.seed(42)


# ================================================================
# MODULE 1: SAMPLE DATA GENERATOR
# In production: replace with pd.read_csv("erp_export.csv")
# ================================================================

def generate_invoice_data(n_normal: int = 500, n_anomalies: int = 30) -> pd.DataFrame:
    """
    Generates realistic invoice data with embedded anomalies.
    In practice, this data comes from SAP, Oracle ERP, or similar systems.

    Args:
        n_normal:    Number of normal (clean) invoices
        n_anomalies: Number of anomalous invoices to embed

    Returns:
        DataFrame with invoice records
    """
    vendors = [
        "Siemens AG", "BASF SE", "Bosch GmbH", "Continental AG",
        "ThyssenKrupp AG", "Henkel AG", "Bayer AG", "Merck KGaA",
        "Fresenius SE", "Deutsche Post AG",
    ]
    categories = [
        "IT-Services", "Consulting", "Raw Materials", "Logistics",
        "Office Supplies", "Marketing", "Maintenance", "Software",
    ]
    cost_centers = ["CC-1001", "CC-1002", "CC-2001", "CC-3001", "CC-4001"]

    start_date = datetime(2024, 1, 1)

    # Realistic amount ranges per category (EUR)
    amount_ranges = {
        "IT-Services":    (5_000,   150_000),
        "Consulting":     (10_000,  300_000),
        "Raw Materials":  (1_000,   500_000),
        "Logistics":      (500,      50_000),
        "Office Supplies":(100,       5_000),
        "Marketing":      (5_000,   200_000),
        "Maintenance":    (1_000,    80_000),
        "Software":       (2_000,   100_000),
    }

    # --- Generate clean invoices ---
    normal_records = []
    for i in range(n_normal):
        date = start_date + timedelta(days=random.randint(0, 364))
        while date.weekday() >= 5:           # Skip weekends for normal invoices
            date += timedelta(days=1)

        category = random.choice(categories)
        lo, hi = amount_ranges[category]

        normal_records.append({
            "invoice_id":  f"INV-2024-{i+1:05d}",
            "date":        date,
            "vendor":      random.choice(vendors),
            "category":    category,
            "amount":      round(random.uniform(lo, hi), 2),
            "currency":    "EUR",
            "cost_center": random.choice(cost_centers),
            "approved_by": f"mgr_{random.randint(1, 10):02d}",
        })

    # --- Embed Anomaly Type 1: Duplicate Invoice IDs ---
    anomaly_records = []
    for _ in range(10):
        original = random.choice(normal_records)
        dup = original.copy()
        # Same invoice ID, slightly different date -> double payment scenario
        dup["date"] = original["date"] + timedelta(days=random.randint(1, 30))
        anomaly_records.append(dup)

    # --- Embed Anomaly Type 2: Extreme Amount Outliers ---
    for j in range(8):
        date = start_date + timedelta(days=random.randint(0, 364))
        while date.weekday() >= 5:
            date += timedelta(days=1)
        anomaly_records.append({
            "invoice_id":  f"INV-ANOM-{j+1:04d}",
            "date":        date,
            "vendor":      random.choice(vendors),
            "category":    random.choice(categories),
            "amount":      round(random.uniform(900_000, 2_000_000), 2),
            "currency":    "EUR",
            "cost_center": random.choice(cost_centers),
            "approved_by": f"mgr_{random.randint(1, 10):02d}",
        })

    # --- Embed Anomaly Type 3: Weekend Bookings ---
    for j in range(7):
        date = start_date + timedelta(days=random.randint(0, 364))
        while date.weekday() < 5:            # Force weekend
            date += timedelta(days=1)
        anomaly_records.append({
            "invoice_id":  f"INV-WE-{j+1:04d}",
            "date":        date,
            "vendor":      random.choice(vendors),
            "category":    random.choice(categories),
            "amount":      round(random.uniform(5_000, 100_000), 2),
            "currency":    "EUR",
            "cost_center": random.choice(cost_centers),
            "approved_by": f"mgr_{random.randint(1, 10):02d}",
        })

    # --- Embed Anomaly Type 4: Round Number Bias (Fraud Indicator) ---
    round_amounts = [50_000.00, 100_000.00, 200_000.00, 75_000.00, 150_000.00]
    for j in range(5):
        date = start_date + timedelta(days=random.randint(0, 364))
        while date.weekday() >= 5:
            date += timedelta(days=1)
        anomaly_records.append({
            "invoice_id":  f"INV-RND-{j+1:04d}",
            "date":        date,
            "vendor":      random.choice(vendors),
            "category":    random.choice(categories),
            "amount":      round_amounts[j],
            "currency":    "EUR",
            "cost_center": random.choice(cost_centers),
            "approved_by": f"mgr_{random.randint(1, 10):02d}",
        })

    all_records = normal_records + anomaly_records
    random.shuffle(all_records)

    df = pd.DataFrame(all_records).reset_index(drop=True)
    df["date"] = pd.to_datetime(df["date"])
    return df


# ================================================================
# MODULE 2: ANOMALY DETECTION ENGINE
# ================================================================

class InvoiceAnomalyDetector:
    """
    Multi-layer anomaly detection engine for invoice data.

    Design principle (how a consultant would approach this):
        - Rule-based checks: fast, explainable, directly actionable
        - ML-based detection: finds unknown patterns without explicit rules
        - Both layers combined -> higher recall, fewer false negatives

    Attributes:
        contamination: Expected anomaly fraction (calibrate with Finance team)
    """

    def __init__(self, contamination: float = 0.05):
        self.contamination = contamination
        self.summary: dict = {}

    def _tag(self, df: pd.DataFrame, anomaly_type: str, risk: str, recommendation: str) -> pd.DataFrame:
        """Helper: adds detection metadata columns to a flagged subset."""
        result = df.copy()
        result["detected_anomaly"] = anomaly_type
        result["risk_level"] = risk
        result["recommendation"] = recommendation
        return result

    def detect_duplicates(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        Flags invoices with duplicate IDs.
        Root cause: same invoice submitted twice or ERP import error.
        Financial impact: direct double payment.
        """
        dupes = df[df.duplicated(subset=["invoice_id"], keep=False)]
        return self._tag(dupes, "Duplicate Invoice ID", "HIGH",
                         "Block payment — investigate with AP team immediately")

    def detect_weekend_bookings(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        Flags invoices booked on Saturdays or Sundays.
        Policy violation: finance staff should not process payments on weekends.
        """
        weekends = df[df["date"].dt.weekday >= 5]
        return self._tag(weekends, "Weekend Booking", "MEDIUM",
                         "Verify manager authorization — escalate if missing")

    def detect_ml_outliers(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        Uses IsolationForest to detect statistical amount outliers.

        Why IsolationForest over Z-Score?
            Invoice amounts are NOT normally distributed — they follow a
            power-law distribution. IsolationForest is robust to this.
            It isolates anomalies by randomly partitioning features and
            measuring how few splits are needed to isolate a point.
        """
        le = LabelEncoder()
        df_enc = df.copy()
        df_enc["vendor_enc"]   = le.fit_transform(df_enc["vendor"])
        df_enc["category_enc"] = le.fit_transform(df_enc["category"])

        features = df_enc[["amount", "vendor_enc", "category_enc"]].values

        clf = IsolationForest(
            contamination=self.contamination,
            n_estimators=100,
            random_state=42,
        )
        predictions = clf.fit_predict(features)

        outliers = df[predictions == -1]
        return self._tag(outliers, "ML Amount Outlier", "HIGH",
                         "Manual review required — verify with vendor and category benchmark")

    def detect_round_numbers(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        Flags high-value invoices with suspiciously round amounts.

        Forensic accounting principle: people fabricating invoices tend
        to use round numbers (50,000 vs. 49,872.43). This is a well-known
        indicator in Benford's Law analysis and forensic audits.
        """
        high_value    = df[df["amount"] >= 10_000]
        round_numbers = high_value[high_value["amount"] % 1_000 == 0]
        return self._tag(round_numbers, "Round Number Bias", "LOW",
                         "Low-priority flag — spot-check supporting documentation")

    def run(self, df: pd.DataFrame) -> dict:
        """
        Executes all detection modules and consolidates findings.

        Args:
            df: Raw invoice DataFrame

        Returns:
            dict with 'all_anomalies', per-type subsets, and 'summary'
        """
        print()
        print("=" * 60)
        print("  INVOICE ANOMALY DETECTOR  |  ANALYSIS RUNNING")
        print("=" * 60)
        print(f"  Invoices analyzed : {len(df):,}")
        print(f"  Date range        : {df['date'].min().date()} -> {df['date'].max().date()}")
        print(f"  Total volume      : EUR {df['amount'].sum():>15,.2f}")
        print("=" * 60)

        duplicates  = self.detect_duplicates(df)
        weekends    = self.detect_weekend_bookings(df)
        ml_outliers = self.detect_ml_outliers(df)
        round_nums  = self.detect_round_numbers(df)

        all_anomalies = (
            pd.concat([duplicates, weekends, ml_outliers, round_nums])
            .drop_duplicates(subset=["invoice_id", "detected_anomaly"])
        )

        high_risk = all_anomalies[all_anomalies["risk_level"] == "HIGH"]

        self.summary = {
            "total_invoices":    len(df),
            "total_anomalies":   len(all_anomalies),
            "anomaly_rate_pct":  round(len(all_anomalies) / len(df) * 100, 1),
            "total_volume_eur":  df["amount"].sum(),
            "anomaly_value_eur": all_anomalies["amount"].sum(),
            "high_risk_count":   len(high_risk),
            "breakdown": {
                "Duplicate Invoice ID": len(duplicates),
                "Weekend Booking":      len(weekends),
                "ML Amount Outlier":    len(ml_outliers),
                "Round Number Bias":    len(round_nums),
            },
        }

        print()
        print(f"  Anomalies found   : {self.summary['total_anomalies']} ({self.summary['anomaly_rate_pct']}%)")
        print(f"  HIGH risk         : {self.summary['high_risk_count']}")
        print(f"  Value at risk     : EUR {self.summary['anomaly_value_eur']:>15,.2f}")
        print()
        for label, count in self.summary["breakdown"].items():
            bar = "#" * count
            print(f"  {label:<26} {count:>3}  {bar}")
        print("=" * 60)
        print()

        return {
            "all_anomalies": all_anomalies,
            "duplicates":    duplicates,
            "weekends":      weekends,
            "ml_outliers":   ml_outliers,
            "round_numbers": round_nums,
            "summary":       self.summary,
        }


# ================================================================
# MODULE 3: EXCEL REPORT GENERATOR
# Typical consulting deliverable: ready-to-share with CFO / audit
# ================================================================

def generate_excel_report(results: dict, output_path: str = "output/anomaly_report.xlsx") -> str:
    """
    Generates a formatted Excel report with executive summary and findings.

    Args:
        results:     Output dict from InvoiceAnomalyDetector.run()
        output_path: File path for the Excel file

    Returns:
        Absolute path to the generated file
    """
    os.makedirs(os.path.dirname(output_path), exist_ok=True)

    # Color palette
    BLUE_HEADER = "002F75B5"
    WHITE       = "00FFFFFF"
    LIGHT_GRAY  = "00F2F2F2"
    RED         = "00FF4444"
    ORANGE      = "00FFA500"
    YELLOW      = "00FFD700"

    wb = Workbook()
    summary = results["summary"]

    # ── Sheet 1: Executive Summary ──────────────────────────────
    ws = wb.active
    ws.title = "Executive Summary"

    def header_cell(cell_ref, text, size=12):
        cell = ws[cell_ref]
        cell.value = text
        cell.font = Font(bold=True, size=size, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=BLUE_HEADER)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        return cell

    header_cell("A1", "INVOICE ANOMALY DETECTION — EXECUTIVE SUMMARY", size=14)
    ws.merge_cells("A1:C1")
    ws.row_dimensions[1].height = 30

    ws["A2"] = f"Report generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws["A2"].font = Font(italic=True, color="00888888")
    ws.merge_cells("A2:C2")

    kpis = [
        ("Total Invoices Analyzed",  f"{summary['total_invoices']:,}"),
        ("Anomalies Detected",        f"{summary['total_anomalies']:,}  ({summary['anomaly_rate_pct']}%)"),
        ("HIGH Risk Findings",        f"{summary['high_risk_count']:,}"),
        ("Total Invoice Volume",      f"EUR {summary['total_volume_eur']:,.2f}"),
        ("Value at Risk (flagged)",   f"EUR {summary['anomaly_value_eur']:,.2f}"),
    ]

    ws["A4"] = "KEY METRICS"
    ws["A4"].font = Font(bold=True, size=11, color=WHITE)
    ws["A4"].fill = PatternFill("solid", fgColor=BLUE_HEADER)
    ws.merge_cells("A4:C4")

    for i, (label, value) in enumerate(kpis, start=5):
        ws[f"A{i}"] = label
        ws[f"B{i}"] = value
        ws[f"A{i}"].font = Font(bold=True)
        if i % 2 == 0:
            for col in ["A", "B", "C"]:
                ws[f"{col}{i}"].fill = PatternFill("solid", fgColor=LIGHT_GRAY)

    ws["A11"] = "ANOMALY BREAKDOWN"
    ws["A11"].font = Font(bold=True, size=11, color=WHITE)
    ws["A11"].fill = PatternFill("solid", fgColor=BLUE_HEADER)
    ws["B11"].fill = PatternFill("solid", fgColor=BLUE_HEADER)
    ws.merge_cells("A11:C11")

    for i, (atype, count) in enumerate(summary["breakdown"].items(), start=12):
        ws[f"A{i}"] = atype
        ws[f"B{i}"] = count
        if i % 2 == 0:
            for col in ["A", "B"]:
                ws[f"{col}{i}"].fill = PatternFill("solid", fgColor=LIGHT_GRAY)

    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 18

    # ── Sheet 2: All Flagged Invoices ───────────────────────────
    ws2 = wb.create_sheet("Flagged Invoices")

    cols = [
        "invoice_id", "date", "vendor", "category", "amount",
        "cost_center", "approved_by", "detected_anomaly", "risk_level", "recommendation",
    ]
    col_labels = [
        "Invoice ID", "Date", "Vendor", "Category", "Amount (EUR)",
        "Cost Center", "Approved By", "Anomaly Type", "Risk Level", "Recommendation",
    ]
    col_widths = [18, 12, 22, 18, 16, 14, 14, 25, 12, 50]
    risk_colors = {"HIGH": RED, "MEDIUM": ORANGE, "LOW": YELLOW}

    for col_idx, label in enumerate(col_labels, 1):
        cell = ws2.cell(row=1, column=col_idx, value=label)
        cell.font = Font(bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=BLUE_HEADER)
        cell.alignment = Alignment(horizontal="center")

    anomaly_df = (
        results["all_anomalies"][cols]
        .copy()
        .sort_values("risk_level")
    )
    anomaly_df["date"] = anomaly_df["date"].dt.strftime("%Y-%m-%d")

    for row_idx, row in enumerate(anomaly_df.itertuples(index=False), start=2):
        for col_idx, value in enumerate(row, start=1):
            cell = ws2.cell(row=row_idx, column=col_idx, value=value)
            # Color-code the Risk Level column
            if col_idx == 9:
                color = risk_colors.get(str(value), WHITE)
                cell.fill = PatternFill("solid", fgColor=color)
                cell.font = Font(bold=True)

    for i, width in enumerate(col_widths, 1):
        ws2.column_dimensions[chr(64 + i)].width = width

    # ── Sheet 3: Raw Data ────────────────────────────────────────
    ws3 = wb.create_sheet("Raw Data")
    raw_cols = ["invoice_id", "date", "vendor", "category", "amount", "currency", "cost_center", "approved_by"]
    raw_df = results["all_anomalies"][raw_cols].copy()  # just anomalies for brevity
    raw_df["date"] = raw_df["date"].dt.strftime("%Y-%m-%d")

    for col_idx, label in enumerate(raw_cols, 1):
        cell = ws3.cell(row=1, column=col_idx, value=label.replace("_", " ").title())
        cell.font = Font(bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=BLUE_HEADER)

    for row_idx, row in enumerate(raw_df.itertuples(index=False), start=2):
        for col_idx, value in enumerate(row, start=1):
            ws3.cell(row=row_idx, column=col_idx, value=value)

    wb.save(output_path)
    abs_path = os.path.abspath(output_path)
    print(f"  Excel report -> {abs_path}")
    return abs_path


# ================================================================
# MODULE 4: VISUALIZATION DASHBOARD
# ================================================================

def generate_charts(results: dict, output_path: str = "output/dashboard.png") -> str:
    """
    Generates a 3-panel dashboard chart for presentations.

    Args:
        results:     Output dict from InvoiceAnomalyDetector.run()
        output_path: File path for the PNG file

    Returns:
        Absolute path to the saved chart
    """
    os.makedirs(os.path.dirname(output_path), exist_ok=True)

    summary = results["summary"]
    breakdown = summary["breakdown"]

    fig, axes = plt.subplots(1, 3, figsize=(17, 5))
    fig.patch.set_facecolor("#F8F9FA")
    fig.suptitle(
        "Invoice Anomaly Detection — Analysis Dashboard",
        fontsize=15, fontweight="bold", y=1.02,
    )

    COLORS = {
        "Duplicate Invoice ID": "#E74C3C",
        "Weekend Booking":      "#E67E22",
        "ML Amount Outlier":    "#C0392B",
        "Round Number Bias":    "#F1C40F",
    }
    PALETTE = [COLORS[k] for k in breakdown.keys()]

    # Chart 1: Anomaly counts by type
    ax1 = axes[0]
    labels = [k.replace(" ", "\n") for k in breakdown.keys()]
    bars = ax1.bar(labels, breakdown.values(), color=PALETTE, edgecolor="white", linewidth=1.5)
    ax1.set_title("Anomalies by Type", fontweight="bold", pad=12)
    ax1.set_ylabel("Count", labelpad=8)
    ax1.set_facecolor("#FFFFFF")
    for bar, val in zip(bars, breakdown.values()):
        ax1.text(
            bar.get_x() + bar.get_width() / 2,
            bar.get_height() + 0.3,
            str(val), ha="center", va="bottom", fontweight="bold", fontsize=11,
        )
    ax1.spines[["top", "right"]].set_visible(False)
    ax1.tick_params(labelsize=8)

    # Chart 2: Normal vs Anomaly pie
    ax2 = axes[1]
    normal_count = summary["total_invoices"] - summary["total_anomalies"]
    ax2.pie(
        [normal_count, summary["total_anomalies"]],
        labels=["Normal", "Anomaly"],
        colors=["#2ECC71", "#E74C3C"],
        autopct="%1.1f%%",
        startangle=90,
        wedgeprops={"edgecolor": "white", "linewidth": 2.5},
        textprops={"fontsize": 11},
    )
    ax2.set_title(
        f"Invoice Health\n({summary['total_invoices']:,} invoices analyzed)",
        fontweight="bold", pad=12,
    )

    # Chart 3: Risk level distribution
    ax3 = axes[2]
    all_anomalies = results["all_anomalies"]
    risk_counts = all_anomalies["risk_level"].value_counts().reindex(["HIGH", "MEDIUM", "LOW"]).fillna(0)
    risk_palette = {"HIGH": "#E74C3C", "MEDIUM": "#F39C12", "LOW": "#F1C40F"}
    bars3 = ax3.bar(
        risk_counts.index,
        risk_counts.values,
        color=[risk_palette[r] for r in risk_counts.index],
        edgecolor="white",
        linewidth=1.5,
    )
    ax3.set_title("Findings by Risk Level", fontweight="bold", pad=12)
    ax3.set_ylabel("Count", labelpad=8)
    ax3.set_facecolor("#FFFFFF")
    for bar, val in zip(bars3, risk_counts.values):
        if val > 0:
            ax3.text(
                bar.get_x() + bar.get_width() / 2,
                bar.get_height() + 0.3,
                int(val), ha="center", va="bottom", fontweight="bold", fontsize=11,
            )
    ax3.spines[["top", "right"]].set_visible(False)

    plt.tight_layout()
    plt.savefig(output_path, dpi=150, bbox_inches="tight", facecolor=fig.get_facecolor())
    plt.close()

    abs_path = os.path.abspath(output_path)
    print(f"  Dashboard chart -> {abs_path}")
    return abs_path


# ================================================================
# MAIN ENTRY POINT
# ================================================================

if __name__ == "__main__":
    # ── Step 1: Load data ───────────────────────────────────────
    # In production: df = pd.read_csv("erp_export.csv", parse_dates=["date"])
    data_path = "data/invoices_sample.csv"
    os.makedirs("data", exist_ok=True)

    print("Generating sample invoice data...")
    df = generate_invoice_data(n_normal=500, n_anomalies=30)
    df.to_csv(data_path, index=False)
    print(f"  Sample data saved -> {os.path.abspath(data_path)}")

    # ── Step 2: Run anomaly detection ───────────────────────────
    detector = InvoiceAnomalyDetector(contamination=0.05)
    results = detector.run(df)

    # ── Step 3: Generate outputs ─────────────────────────────────
    print("[OUTPUT GENERATION]")
    generate_excel_report(results)
    generate_charts(results)

    print()
    print("  All outputs saved to /output/")
    print("  Open anomaly_report.xlsx for the full Excel report.")
    print()
